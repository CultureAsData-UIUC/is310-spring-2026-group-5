import csv
from openpyxl import load_workbook

excel_file = "ai_k12_education_articles.xlsx"
csv_file = "ai_k12_education_articles.csv"

workbook = load_workbook(excel_file)
sheet = workbook.active

with open(csv_file, "w", newline="", encoding="utf-8-sig") as outfile:
    writer = csv.writer(outfile, quoting=csv.QUOTE_ALL)

    for row in sheet.iter_rows(values_only=True):
        cleaned_row = [
            "" if cell is None else str(cell).strip()
            for cell in row
        ]
        writer.writerow(cleaned_row)

print("CSV created successfully from Excel file.")